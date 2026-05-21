"""
excel_report.py
Generates a formatted multi-sheet Excel workbook with KPI summaries,
pivot tables, charts, and conditional formatting — ready for business stakeholders.
"""

import pandas as pd
import numpy as np
import os

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
        GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

PROCESSED_DIR = "../../data/processed"
EXPORTS_DIR = "../../data/exports"
OUTPUT_PATH = "../../excel/Customer_Behavior_Analytics_Report.xlsx"

os.makedirs("../../excel", exist_ok=True)


# ─────────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────────

DARK_BLUE = "1F3864"
MED_BLUE = "2E75B6"
LIGHT_BLUE = "DEEAF1"
WHITE = "FFFFFF"
ORANGE = "ED7D31"
GREEN = "70AD47"
RED = "FF0000"
GREY = "F2F2F2"


def hdr_style(cell, bg=DARK_BLUE, fg=WHITE, bold=True, size=11, wrap=False):
    cell.font = Font(bold=bold, color=fg, size=size, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def title_style(cell, size=14):
    cell.font = Font(bold=True, size=size, color=DARK_BLUE, name="Calibri Light")
    cell.alignment = Alignment(horizontal="left", vertical="center")


def thin_border():
    s = Side(border_style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ─────────────────────────────────────────────
# SHEET 1: KPI DASHBOARD
# ─────────────────────────────────────────────

def add_kpi_sheet(wb, customers, transactions, rfm):
    ws = wb.create_sheet("📊 KPI Dashboard")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("B2:J2")
    ws["B2"] = "CUSTOMER BEHAVIOR ANALYTICS — EXECUTIVE KPI DASHBOARD"
    title_style(ws["B2"], size=16)

    # KPI Cards
    kpis = [
        ("Total Customers", f"{len(customers):,}"),
        ("Total Transactions", f"{len(transactions):,}"),
        ("Total Revenue", f"${transactions['revenue'].sum():,.0f}"),
        ("Avg Order Value", f"${transactions['final_amount'].mean():,.2f}"),
        ("Avg Rating", f"{transactions['rating'].mean():.2f} / 5"),
        ("Churn Rate", f"{customers['is_churned'].mean()*100:.1f}%"),
        ("Return Rate", f"{transactions['is_returned'].mean()*100:.1f}%"),
        ("Active Customers", f"{(customers['is_churned']==0).sum():,}"),
    ]

    col_positions = [2, 4, 6, 8, 2, 4, 6, 8]
    row_positions = [4, 4, 4, 4, 7, 7, 7, 7]

    for i, (label, value) in enumerate(kpis):
        c = col_positions[i]
        r = row_positions[i]
        col_letter = get_column_letter(c)
        next_col = get_column_letter(c + 1)
        ws.merge_cells(f"{col_letter}{r}:{next_col}{r}")
        ws.merge_cells(f"{col_letter}{r+1}:{next_col}{r+1}")
        ws[f"{col_letter}{r}"] = label
        ws[f"{col_letter}{r+1}"] = value
        hdr_style(ws[f"{col_letter}{r}"], bg=MED_BLUE, size=10)
        ws[f"{col_letter}{r+1}"].font = Font(bold=True, size=16, color=DARK_BLUE)
        ws[f"{col_letter}{r+1}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"{col_letter}{r+1}"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)

    ws.row_dimensions[5].height = 30
    ws.row_dimensions[8].height = 30
    set_col_widths(ws, {get_column_letter(i): 16 for i in range(1, 12)})


# ─────────────────────────────────────────────
# SHEET 2: REVENUE ANALYSIS
# ─────────────────────────────────────────────

def add_revenue_sheet(wb, transactions):
    ws = wb.create_sheet("📈 Revenue Analysis")
    ws.sheet_view.showGridLines = False

    monthly = transactions.groupby(
        [transactions["transaction_date"].str[:7]]
    ).agg(
        revenue=("revenue", "sum"),
        transactions=("transaction_id", "count"),
        avg_value=("final_amount", "mean")
    ).reset_index()
    monthly.columns = ["Month", "Revenue", "Transactions", "Avg Order Value"]
    monthly["Revenue"] = monthly["Revenue"].round(2)
    monthly["Avg Order Value"] = monthly["Avg Order Value"].round(2)

    ws["B2"] = "MONTHLY REVENUE ANALYSIS"
    title_style(ws["B2"])

    headers = list(monthly.columns)
    for ci, h in enumerate(headers, start=2):
        c = ws.cell(row=4, column=ci, value=h)
        hdr_style(c)

    for ri, row in monthly.iterrows():
        for ci, val in enumerate(row, start=2):
            cell = ws.cell(row=ri + 5, column=ci, value=val)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center")
            if ci == 3:  # Revenue
                cell.number_format = '"$"#,##0.00'
            if (ri + 5) % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=GREY)

    # Conditional formatting on revenue
    max_row = len(monthly) + 5
    ws.conditional_formatting.add(
        f"C5:C{max_row}",
        ColorScaleRule(start_type="min", start_color="FFFFFF",
                       end_type="max", end_color=MED_BLUE)
    )

    # Line chart
    chart = LineChart()
    chart.title = "Monthly Revenue Trend"
    chart.style = 10
    chart.y_axis.title = "Revenue ($)"
    chart.x_axis.title = "Month"

    data_ref = Reference(ws, min_col=3, min_row=4, max_row=max_row)
    cats_ref = Reference(ws, min_col=2, min_row=5, max_row=max_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.line.solidFill = MED_BLUE
    chart.width = 20
    chart.height = 12
    ws.add_chart(chart, "G5")

    set_col_widths(ws, {"B": 14, "C": 16, "D": 16, "E": 18})


# ─────────────────────────────────────────────
# SHEET 3: CUSTOMER SEGMENTS
# ─────────────────────────────────────────────

def add_segments_sheet(wb, customers):
    ws = wb.create_sheet("👥 Customer Segments")
    ws.sheet_view.showGridLines = False

    ws["B2"] = "CUSTOMER SEGMENT ANALYSIS"
    title_style(ws["B2"])

    seg = customers.groupby("segment").agg(
        customer_count=("customer_id", "count"),
        avg_age=("age", "mean"),
        avg_loyalty=("loyalty_points", "mean"),
        churn_rate=("is_churned", "mean")
    ).reset_index()
    seg["avg_age"] = seg["avg_age"].round(1)
    seg["avg_loyalty"] = seg["avg_loyalty"].round(0)
    seg["churn_rate"] = (seg["churn_rate"] * 100).round(2)
    seg.columns = ["Segment", "Customers", "Avg Age", "Avg Loyalty Pts", "Churn Rate (%)"]

    headers = list(seg.columns)
    for ci, h in enumerate(headers, start=2):
        c = ws.cell(row=4, column=ci, value=h)
        hdr_style(c, bg=ORANGE)

    for ri, row in seg.iterrows():
        for ci, val in enumerate(row, start=2):
            cell = ws.cell(row=ri + 5, column=ci, value=val)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center")

    # Bar chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Customers by Segment"
    chart.y_axis.title = "Count"
    max_row = len(seg) + 5
    data_ref = Reference(ws, min_col=3, min_row=4, max_row=max_row)
    cats_ref = Reference(ws, min_col=2, min_row=5, max_row=max_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.width = 16
    chart.height = 12
    ws.add_chart(chart, "H5")

    set_col_widths(ws, {"B": 16, "C": 14, "D": 10, "E": 18, "F": 18})


# ─────────────────────────────────────────────
# SHEET 4: CATEGORY PERFORMANCE
# ─────────────────────────────────────────────

def add_category_sheet(wb, transactions):
    ws = wb.create_sheet("🛒 Category Performance")
    ws.sheet_view.showGridLines = False

    ws["B2"] = "PRODUCT CATEGORY PERFORMANCE"
    title_style(ws["B2"])

    cat = transactions.groupby("product_category").agg(
        revenue=("revenue", "sum"),
        orders=("transaction_id", "count"),
        avg_value=("final_amount", "mean"),
        avg_rating=("rating", "mean"),
        return_rate=("is_returned", "mean")
    ).reset_index().sort_values("revenue", ascending=False)
    cat["revenue"] = cat["revenue"].round(2)
    cat["avg_value"] = cat["avg_value"].round(2)
    cat["avg_rating"] = cat["avg_rating"].round(2)
    cat["return_rate"] = (cat["return_rate"] * 100).round(2)
    cat.columns = ["Category", "Revenue ($)", "Orders", "Avg Value ($)", "Avg Rating", "Return Rate (%)"]

    headers = list(cat.columns)
    for ci, h in enumerate(headers, start=2):
        c = ws.cell(row=4, column=ci, value=h)
        hdr_style(c, bg=GREEN)

    for ri, row in cat.iterrows():
        for ci, val in enumerate(row, start=2):
            cell = ws.cell(row=ri + 5, column=ci, value=val)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center")
            if (ri + 5) % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=GREY)

    set_col_widths(ws, {"B": 18, "C": 16, "D": 14, "E": 16, "F": 14, "G": 16})


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def generate_excel_report():
    print("=" * 60)
    print("  EXCEL REPORT GENERATOR")
    print("=" * 60)

    if not HAS_OPENPYXL:
        print("  ⚠️  openpyxl not installed. Install with: pip install openpyxl")
        print("  Generating CSV fallback exports instead...")
        customers = pd.read_csv(f"{PROCESSED_DIR}/customers_cleaned.csv")
        transactions = pd.read_csv(f"{PROCESSED_DIR}/transactions_cleaned.csv",
                                   parse_dates=["transaction_date"])
        rfm = pd.read_csv(f"{PROCESSED_DIR}/rfm_scores.csv")

        customers.to_csv(f"{EXPORTS_DIR}/excel_customers.csv", index=False)
        transactions.head(10000).to_csv(f"{EXPORTS_DIR}/excel_transactions_sample.csv", index=False)
        rfm.to_csv(f"{EXPORTS_DIR}/excel_rfm.csv", index=False)
        print("  ✅ CSV fallbacks saved to data/exports/")
        return

    customers = pd.read_csv(f"{PROCESSED_DIR}/customers_cleaned.csv",
                            parse_dates=["join_date"])
    transactions = pd.read_csv(f"{PROCESSED_DIR}/transactions_cleaned.csv",
                               parse_dates=["transaction_date"])
    rfm = pd.read_csv(f"{PROCESSED_DIR}/rfm_scores.csv")

    transactions["transaction_date"] = transactions["transaction_date"].astype(str)

    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    print("  Building sheets...")
    add_kpi_sheet(wb, customers, transactions, rfm)
    add_revenue_sheet(wb, transactions)
    add_segments_sheet(wb, customers)
    add_category_sheet(wb, transactions)

    wb.save(OUTPUT_PATH)
    print(f"\n  ✅ Excel report saved: {OUTPUT_PATH}")


if __name__ == "__main__":
    generate_excel_report()
