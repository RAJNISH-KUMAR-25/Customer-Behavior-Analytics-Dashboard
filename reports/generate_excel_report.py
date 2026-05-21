import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
import random
from datetime import datetime, timedelta

def generate_sample_kpi_data():
    months = pd.date_range("2022-01", periods=24, freq="ME").strftime("%b %Y").tolist()
    revenue = [round(random.uniform(80000, 200000), 2) for _ in months]
    orders  = [random.randint(3000, 8000) for _ in months]
    churn   = [round(random.uniform(3, 12), 2) for _ in months]
    aov     = [round(r / o, 2) for r, o in zip(revenue, orders)]
    return pd.DataFrame({
        "Month": months, "Revenue": revenue,
        "Orders": orders, "Avg_Order_Value": aov, "Churn_Rate_Pct": churn
    })

def style_header(ws, row, cols, color="1F3864"):
    fill = PatternFill("solid", fgColor=color)
    font = Font(bold=True, color="FFFFFF", size=11)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")

def build_report():
    wb = Workbook()

    # ── Sheet 1: KPI Summary ──
    ws1 = wb.active
    ws1.title = "KPI Summary"
    df = generate_sample_kpi_data()

    ws1.append(["Customer Behavior Analytics - KPI Report"])
    ws1["A1"].font = Font(bold=True, size=16, color="1F3864")
    ws1.append([f"Generated: {datetime.now().strftime('%Y-%m-%d')}"])
    ws1.append([])

    headers = ["Month", "Revenue ($)", "Orders", "Avg Order Value ($)", "Churn Rate (%)"]
    ws1.append(headers)
    style_header(ws1, 4, len(headers))

    for _, row in df.iterrows():
        ws1.append([row["Month"], row["Revenue"], row["Orders"],
                    row["Avg_Order_Value"], row["Churn_Rate_Pct"]])

    for col in range(1, 6):
        ws1.column_dimensions[get_column_letter(col)].width = 22

    # Bar Chart – Revenue
    chart = BarChart()
    chart.title = "Monthly Revenue"
    chart.style = 10
    chart.y_axis.title = "Revenue ($)"
    chart.x_axis.title = "Month"
    data = Reference(ws1, min_col=2, min_row=4, max_row=4 + len(df))
    cats = Reference(ws1, min_col=1, min_row=5, max_row=4 + len(df))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    ws1.add_chart(chart, "G4")

    # ── Sheet 2: Customer Segments ──
    ws2 = wb.create_sheet("Customer Segments")
    seg_data = [
        ["Segment", "Customers", "Avg Revenue", "Avg Orders", "Churn Risk"],
        ["Champions",    200, 4500.00, 32, "Low"],
        ["Loyal",        300, 2800.00, 20, "Low"],
        ["At Risk",      250, 1200.00,  8, "High"],
        ["Lost",         150,  400.00,  2, "Very High"],
        ["New",          100,  300.00,  1, "Medium"],
    ]
    for i, row in enumerate(seg_data, 1):
        ws2.append(row)
        if i == 1:
            style_header(ws2, 1, 5)
    for col in range(1, 6):
        ws2.column_dimensions[get_column_letter(col)].width = 20

    # ── Sheet 3: Category Performance ──
    ws3 = wb.create_sheet("Category Performance")
    cat_data = [
        ["Category", "Transactions", "Revenue ($)", "Avg Amount ($)", "Return Rate (%)"],
        ["Electronics", 22000, 4500000, 204.5, 10.2],
        ["Clothing",    25000, 2100000,  84.0,  9.1],
        ["Grocery",     30000, 1800000,  60.0,  3.5],
        ["Home",        15000, 2800000, 186.7,  6.8],
        ["Sports",       8000,  950000, 118.8,  5.4],
    ]
    for i, row in enumerate(cat_data, 1):
        ws3.append(row)
        if i == 1:
            style_header(ws3, 1, 5)
    for col in range(1, 6):
        ws3.column_dimensions[get_column_letter(col)].width = 22

    wb.save("reports/KPI_Report.xlsx")
    print("Excel report saved → reports/KPI_Report.xlsx")

if __name__ == "__main__":
    build_report()
